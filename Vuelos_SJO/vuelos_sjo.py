#!/usr/bin/env python3
"""
Vuelos SJO - Python Version
============================
Procesa datos de vuelos del Aeropuerto Juan Santamaría (SJO),
filtra vuelos internacionales excluyendo rutas domésticas y ciertas
aerolíneas, y genera un reporte organizado por bloques horarios.

Uso:
    python3 vuelos_sjo.py vuelos.xlsx [reporte.xlsx]

Columnas esperadas en el origen:
    A: Hora | B: N° Vuelo | C: Destino | D: Aerolínea

Requisitos: openpyxl
"""

import sys
import os
from datetime import datetime, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Configuración ──
EXCLUIR_DESTINOS = [
    "cobano", "tamarindo", "puerto jimenez", "limon", "golfito",
    "liberia", "quepos", "drake bay", "nosara", "la fortuna",
    "bocas del toro", "managua", "bocas", "tortuguero",
]

EXCLUIR_AEROLINEAS = [
    "sansa", "latam", "getjet", "netjets", "viva", "mas air",
]

BLOQUES = [
    ("05 a 09", time(5, 0), time(8, 59)),
    ("09 a 12", time(9, 0), time(11, 59)),
    ("12 a 15", time(12, 0), time(14, 59)),
    ("15 a 18", time(15, 0), time(17, 59)),
    ("18 a 21", time(18, 0), time(20, 59)),
    ("21 a 05", time(21, 0), time(23, 59)),
]

COLOR_PRIMARY = "1F4E79"
COLOR_SECONDARY = "D6E4F0"
COLOR_BG_LIGHT = "DEEAF1"
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"

FILL_PRIMARY = PatternFill("solid", fgColor=COLOR_PRIMARY)
FILL_SECONDARY = PatternFill("solid", fgColor=COLOR_SECONDARY)
FILL_ZEBRA1 = PatternFill("solid", fgColor=COLOR_BG_LIGHT)
FILL_ZEBRA2 = PatternFill("solid", fgColor=COLOR_WHITE)
FILL_BLOQUE = PatternFill("solid", fgColor="BDD7EE")

FONT_TITLE = Font(name="Arial", bold=True, size=13, color=COLOR_WHITE)
FONT_DATE = Font(name="Arial", bold=True, size=11, color=COLOR_PRIMARY)
FONT_HEADER = Font(name="Arial", bold=True, size=10, color=COLOR_WHITE)
FONT_BLOQUE = Font(name="Arial", bold=True, size=10, color=COLOR_PRIMARY)
FONT_DATA = Font(name="Arial", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def parse_hora(val):
    """Convierte un valor de celda a objeto time."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, time):
        return val
    try:
        # string format
        s = str(val).strip()
        for fmt in ["%H:%M", "%H:%M:%S", "%I:%M %p", "%H.%M"]:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.time()
            except ValueError:
                continue
    except Exception:
        pass
    return None


def en_bloque(hora_time, ini, fin, es_ultimo):
    """Verifica si la hora está en el bloque."""
    if hora_time is None:
        return False
    if es_ultimo:
        return hora_time >= ini  # 21:00 en adelante
    return ini <= hora_time <= fin


def normalizar_destino(raw):
    """Limpia el nombre del destino."""
    s = str(raw or "").strip()
    # Quitar paréntesis y contenido
    idx = s.find("(")
    if idx > 0:
        s = s[:idx].strip()
    return s.replace("\u00a0", "").strip()


def debe_excluir(destino, aerolinea):
    """Determina si el vuelo debe excluirse."""
    d = destino.lower()
    a = aerolinea.lower()
    for ex in EXCLUIR_DESTINOS:
        if ex in d:
            return True
    for ex in EXCLUIR_AEROLINEAS:
        if ex in a:
            return True
    return False


def process(file_in, file_out="Reporte_Vuelos_SJO.xlsx"):
    wb_src = load_workbook(file_in, data_only=True)
    ws_src = wb_src.active

    vuelos = []

    for row in ws_src.iter_rows(min_row=1, values_only=True):
        if len(row) < 4:
            continue
        hora = parse_hora(row[0])
        if hora is None or hora == time(0, 0):
            continue

        destino = normalizar_destino(row[2])
        aerolinea = str(row[3] or "").strip().replace("\u00a0", "").strip()
        nro_vuelo = str(row[1] or "").strip() if len(row) > 1 else ""

        if not aerolinea or aerolinea == "-":
            continue
        if debe_excluir(destino, aerolinea):
            continue

        vuelos.append({
            "hora": hora,
            "nro_vuelo": nro_vuelo,
            "destino": destino,
            "aerolinea": aerolinea,
        })

    if not vuelos:
        print("No se encontraron vuelos internacionales.")
        return

    # Crear libro
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Anchos de columna
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10

    # ── Título ──
    ws.merge_cells("A1:C1")
    c1 = ws.cell(row=1, column=1, value="Vuelos del D\u00eda - Aeropuerto Juan Santamar\u00eda")
    c1.font = FONT_TITLE
    c1.fill = FILL_PRIMARY
    c1.alignment = CENTER
    c1.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # ── Fecha ──
    ws.merge_cells("A2:C2")
    ahora = datetime.now()
    dias_es = ["lunes", "martes", "mi\u00e9rcoles", "jueves", "viernes", "s\u00e1bado", "domingo"]
    meses_es = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                "julio", "agosto", "setiembre", "octubre", "noviembre", "diciembre"]
    fecha_str = f"{dias_es[ahora.weekday()]} {ahora.day} de {meses_es[ahora.month-1]} de {ahora.year}"
    c2 = ws.cell(row=2, column=1, value=fecha_str)
    c2.font = FONT_DATE
    c2.fill = FILL_SECONDARY
    c2.alignment = CENTER
    c2.border = THIN_BORDER
    ws.row_dimensions[2].height = 20

    # ── Encabezados ──
    encs = ["Aerol\u00ednea", "Destino", "Hora"]
    for ci, h in enumerate(encs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_PRIMARY
        c.alignment = CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[3].height = 18

    # ── Bloques horarios ──
    fila = 4
    total_vuelos = 0

    for bloque_nombre, ini, fin in BLOQUES:
        es_ultimo = (bloque_nombre == "21 a 05")
        vuelos_bloque = [v for v in vuelos if en_bloque(v["hora"], ini, fin, es_ultimo)]
        if not vuelos_bloque:
            continue

        # Título del bloque
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        cb = ws.cell(row=fila, column=1, value=bloque_nombre)
        cb.font = FONT_BLOQUE
        cb.fill = FILL_BLOQUE
        cb.alignment = CENTER
        cb.border = THIN_BORDER
        ws.row_dimensions[fila].height = 16
        fila += 1

        # Vuelos en el bloque
        for vi, v in enumerate(vuelos_bloque):
            fill = FILL_ZEBRA1 if vi % 2 == 0 else FILL_ZEBRA2
            ws.cell(row=fila, column=1, value=v["aerolinea"]).font = FONT_DATA
            ws.cell(row=fila, column=1).fill = fill
            ws.cell(row=fila, column=1).alignment = LEFT
            ws.cell(row=fila, column=2, value=v["destino"]).font = FONT_DATA
            ws.cell(row=fila, column=2).fill = fill
            ws.cell(row=fila, column=2).alignment = LEFT
            ws.cell(row=fila, column=3, value=v["hora"].strftime("%H:%M")).font = FONT_DATA
            ws.cell(row=fila, column=3).fill = fill
            ws.cell(row=fila, column=3).alignment = CENTER

            for col in range(1, 4):
                ws.cell(row=fila, column=col).border = THIN_BORDER
            ws.row_dimensions[fila].height = 15
            total_vuelos += 1
            fila += 1

    wb.save(file_out)
    print(f"\u2705 Reporte de vuelos guardado: {file_out} ({total_vuelos} vuelos)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 vuelos_sjo.py vuelos.xlsx [reporte_salida.xlsx]")
        sys.exit(1)
    archivo_in = sys.argv[1]
    archivo_out = sys.argv[2] if len(sys.argv) > 2 else "Reporte_Vuelos_SJO.xlsx"
    process(archivo_in, archivo_out)

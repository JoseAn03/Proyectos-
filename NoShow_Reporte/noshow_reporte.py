#!/usr/bin/env python3
"""
NoShow Reporte - Python Version
================================
Genera un reporte de No Show con todas las reservas activas del día,
ordenadas alfabéticamente por nombre del cliente.

Uso:
    python3 noshow_reporte.py manifiesto.xlsx [salida.xlsx]

Requisitos: openpyxl
"""

import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Constantes ──
COLOR_AZUL = "0070C0"
COLOR_NEGRO = "000000"
COLOR_VERDE = "00B050"
COLOR_ROJO = "FF0000"
COLOR_HEADER_BG = "1F4E79"
COLOR_BLANCO = "FFFFFF"

ESTRELLA = "\u2B50"

HEADERS = [
    "Letra", "N.\u00ba reserva", "Marca", "**", "Nombre", "Clase",
    "**MAIN VIP**", "Fecha recogida", "Fecha entrega",
    "Ubic. rec.", "Ubic. ent.", "Tarifa Diaria", "Agencia de recom."
]

FONT_HEADER = Font(name="Calibri", bold=True, size=11, color=COLOR_BLANCO)
FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER_BG)
THIN_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)


def determinar_marca(ubic):
    """Retorna (nombre_marca, letra_marca, color_hex)."""
    u = ubic.upper() if ubic else ""
    if "E71" in u or "T71" in u:
        return "ALAMO/", "I", COLOR_AZUL
    elif "C61" in u or "T61" in u:
        return "ENTERPRISE/", "P", COLOR_NEGRO
    elif "E01" in u or "T01" in u:
        return "NATIONAL/", "E", COLOR_VERDE
    return "", "", COLOR_NEGRO


def calc_columna_d(flag_a, plus_c, letra_marca):
    a = str(flag_a or "").strip()
    c = str(plus_c or "").strip()
    has_fl = ("~" in a or "*" in a)
    has_plus = ("+" in c)
    if has_fl and has_plus and letra_marca:
        return f"{ESTRELLA}{letra_marca}-FL{ESTRELLA}"
    elif has_fl:
        return f"{ESTRELLA}FL{ESTRELLA}"
    elif has_plus and letra_marca:
        return letra_marca
    return ""


def calc_main_vip(ubic, agencia):
    u = (ubic or "").upper().strip()
    a = str(agencia or "").strip()
    vip = u.endswith("E71") or u.endswith("E01") or u.endswith("C61")
    exp = "11617270" in a
    if vip and exp:
        return "* MAIN VIP* * EXPEDIA*", "both"
    elif vip:
        return "* MAIN VIP*", "vip"
    elif exp:
        return "* EXPEDIA*", "expedia"
    return "", "none"


def process(file_in, file_out="NoShow_Reporte.xlsx"):
    wb_src = load_workbook(file_in, data_only=True)
    ws_src = wb_src.active

    registros = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if len(row) < 18:
            continue
        ubic = str(row[9] or "").strip()
        if not ubic:
            continue
        marca, letra, color = determinar_marca(ubic)
        if not marca:
            continue

        nombre = str(row[3] or "").strip()
        if not nombre:
            continue

        registros.append({
            "flag_a": row[0],
            "nro_reserva": row[1],
            "plus_c": row[2],
            "nombre": nombre,
            "clase": row[5],
            "f_rec": row[7],
            "f_ent": row[8],
            "ubic_rec": row[9],
            "ubic_ent": row[11],
            "tarifa": row[12],
            "agencia": row[17],
            "marca": marca,
            "letra_marca": letra,
            "color": color,
        })

    if not registros:
        print("No se encontraron datos.")
        return

    # Ordenar por nombre
    registros.sort(key=lambda r: r["nombre"].lower())

    # Crear libro
    wb = Workbook()
    ws = wb.active
    ws.title = "Manifiesto"

    # Encabezados
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[1].height = 28

    widths = [7, 14, 13, 11, 28, 7, 20, 18, 18, 11, 11, 11, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    letra_ant = ""
    for fi, r in enumerate(registros, 2):
        pc = (r["nombre"] or "").strip()
        lc = pc[0].upper() if pc and "A" <= pc[0].upper() <= "Z" else ""
        letra_val = lc if lc and lc != letra_ant else ""
        letra_ant = lc if lc else letra_ant

        from datetime import datetime
        f_rec = ""
        if isinstance(r["f_rec"], datetime):
            f_rec = r["f_rec"].strftime("%d/%m/%Y %H:%M")
        elif r["f_rec"]:
            f_rec = str(r["f_rec"])
        f_ent = ""
        if isinstance(r["f_ent"], datetime):
            f_ent = r["f_ent"].strftime("%d/%m/%Y %H:%M")
        elif r["f_ent"]:
            f_ent = str(r["f_ent"])

        col_d = calc_columna_d(r["flag_a"], r["plus_c"], r["letra_marca"])
        col_g, vip_tipo = calc_main_vip(r["ubic_rec"], r["agencia"])

        ws.cell(row=fi, column=1, value=letra_val)
        ws.cell(row=fi, column=2, value=r["nro_reserva"])
        ws.cell(row=fi, column=3, value=r["marca"])
        ws.cell(row=fi, column=4, value=col_d)
        ws.cell(row=fi, column=5, value=r["nombre"])
        ws.cell(row=fi, column=6, value=r["clase"])
        ws.cell(row=fi, column=7, value=col_g)
        ws.cell(row=fi, column=8, value=f_rec)
        ws.cell(row=fi, column=9, value=f_ent)
        ws.cell(row=fi, column=10, value=r["ubic_rec"])
        ws.cell(row=fi, column=11, value=r["ubic_ent"])
        ws.cell(row=fi, column=12, value=r["tarifa"])
        ws.cell(row=fi, column=13, value=r["agencia"])

        b_font = Font(name="Calibri", bold=True, size=10)
        for col in range(1, 14):
            c = ws.cell(row=fi, column=col)
            c.font = b_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER

        if letra_val:
            ws.cell(row=fi, column=1).font = Font(name="Calibri", bold=True, size=11, color=COLOR_ROJO)
            ws.cell(row=fi, column=1).fill = PatternFill("solid", fgColor="FFFF00")

        ws.cell(row=fi, column=3).font = Font(name="Calibri", bold=True, size=10, color=r["color"])
        if col_d:
            ws.cell(row=fi, column=4).font = Font(name="Calibri", bold=True, size=10, color=COLOR_ROJO)
        ws.cell(row=fi, column=5).font = Font(name="Calibri", bold=True, size=10, color=r["color"])

        if vip_tipo == "vip":
            ws.cell(row=fi, column=7).fill = PatternFill("solid", fgColor="C6EFCE")
        elif vip_tipo == "expedia":
            ws.cell(row=fi, column=7).fill = PatternFill("solid", fgColor="FFC7CE")
        elif vip_tipo == "both":
            ws.cell(row=fi, column=7).fill = PatternFill("solid", fgColor="FFD966")

        if letra_val:
            for col in range(1, 14):
                c2 = ws.cell(row=fi, column=col)
                c2.border = Border(
                    left=Side(style="thin", color="AAAAAA"),
                    right=Side(style="thin", color="AAAAAA"),
                    top=Side(style="medium", color=COLOR_NEGRO),
                    bottom=Side(style="thin", color="AAAAAA"),
                )

        ws.row_dimensions[fi].height = 18

    ws.freeze_panes = "A2"
    wb.save(file_out)
    print(f"✅ Reporte NoShow guardado: {file_out} ({len(registros)} registros)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 noshow_reporte.py manifiesto.xlsx [salida.xlsx]")
        sys.exit(1)
    archivo_in = sys.argv[1]
    archivo_out = sys.argv[2] if len(sys.argv) > 2 else "NoShow_Reporte.xlsx"
    process(archivo_in, archivo_out)

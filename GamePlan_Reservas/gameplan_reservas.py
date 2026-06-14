#!/usr/bin/env python3
"""
GamePlan Reservas - Python Version
===================================
Procesa el manifiesto de reservas (Excel) y genera un nuevo archivo
con tres pestañas (ALAMO, ENTERPRISE, NATIONAL) ordenadas alfabéticamente
por nombre, con formato condicional, colores por marca e indicadores.

Uso:
    python3 gameplan_reservas.py manifiesto.xlsx [salida.xlsx]

Requisitos: openpyxl
"""

import sys
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy

# ── Constantes de estilo ──
COLOR_AZUL = "0070C0"
COLOR_NEGRO = "000000"
COLOR_VERDE = "00B050"
COLOR_ROJO = "FF0000"
COLOR_AMARILLO = "FFFF00"
COLOR_ROJO_CLARO = "FFC7CE"
COLOR_VERDE_FONDO = "C6EFCE"
COLOR_BLANCO = "FFFFFF"
COLOR_HEADER_BG = "1F4E79"

FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_WHITE = Font(name="Calibri", bold=True, size=10, color=COLOR_BLANCO)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER_BG)
FILL_YELLOW = PatternFill("solid", fgColor=COLOR_AMARILLO)
FILL_GREEN = PatternFill("solid", fgColor=COLOR_VERDE_FONDO)
FILL_RED_LIGHT = PatternFill("solid", fgColor=COLOR_ROJO_CLARO)

THIN_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)
TOP_BORDER = Border(
    top=Side(style="medium", color=COLOR_NEGRO),
)

HEADERS = [
    "Letra", "N.\u00ba reserva", "Marca", "**", "Nombre", "Clase",
    "**MAIN VIP**", "Fecha recogida", "Fecha entrega",
    "Ubic. rec.", "Ubic. ent.", "Tarifa Diaria", "Agencia de recom."
]

ESTRELLA = "\u2B50"


def parse_fecha(val):
    """Convierte valor de celda a string de fecha formateada."""
    if val is None:
        return ""
    from datetime import datetime
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    return str(val).strip()


def determinar_marca(ubic_rec):
    """Devuelve (marca_nombre, letra_marca) según la ubicación."""
    u = ubic_rec.upper() if ubic_rec else ""
    if "E71" in u or "T71" in u:
        return "ALAMO/", "I", COLOR_AZUL
    elif "C61" in u or "T61" in u:
        return "ENTERPRISE/", "P", COLOR_NEGRO
    elif "E01" in u or "T01" in u:
        return "NATIONAL/", "E", COLOR_VERDE
    return "", "", COLOR_NEGRO


def calcular_columna_d(valor_col_a, valor_col_c, letra_marca):
    """
    Columna **:
    - + en col C original → I/P/E según marca
    - ~ o * en col A original → ⭐FL⭐
    - Ambos → combinado ⭐I-FL⭐, ⭐P-FL⭐, ⭐E-FL⭐
    """
    a = str(valor_col_a or "").strip()
    c = str(valor_col_c or "").strip()
    has_fl = ("~" in a or "*" in a)
    has_plus = ("+" in c)

    if has_fl and has_plus and letra_marca:
        return f"{ESTRELLA}{letra_marca}-FL{ESTRELLA}"
    elif has_fl:
        return f"{ESTRELLA}FL{ESTRELLA}"
    elif has_plus and letra_marca:
        return letra_marca
    return ""


def calcular_main_vip(ubic_rec, agencia):
    """
    Columna **MAIN VIP**:
    - * MAIN VIP* si ubic termina en E71/E01/C61 → fondo verde
    - * EXPEDIA* si agencia contiene 11617270 → fondo rojo claro
    """
    u = (ubic_rec or "").upper().strip()
    a = str(agencia or "").strip()
    is_vip = u.endswith("E71") or u.endswith("E01") or u.endswith("C61")
    is_expedia = "11617270" in a

    if is_vip and is_expedia:
        return "* MAIN VIP* * EXPEDIA*", "both"
    elif is_vip:
        return "* MAIN VIP*", "vip"
    elif is_expedia:
        return "* EXPEDIA*", "expedia"
    return "", "none"


def escribir_hoja(ws, datos, marca_nombre, letra_marca, color_hex):
    """Escribe y formatea una pestaña con los datos filtrados."""
    color_int = int(color_hex, 16)

    # Ordenar por nombre (columna 4 en datos = nombre original, col A=1, B=2, C=3, D=4...)
    # datos[i] = [col_a, col_b(col2), col_c, col_d, col_e, col_f, col_g, col_h,
    #             col_i, col_j, col_k, col_m, col_r(col18),
    #             flag_col_a_orig, plus_col_c_orig]
    datos.sort(key=lambda r: (r[4] or "").lower())

    # Encabezados
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_WHITE
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28

    # Anchos de columna
    widths = [7, 14, 13, 11, 28, 7, 20, 18, 18, 11, 11, 11, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    letra_anterior = ""
    fila = 2

    for d in datos:
        flag_col_a = str(d[13] or "").strip()
        plus_col_c = str(d[14] or "").strip()
        nro_reserva = d[1]
        nombre = d[4]
        clase = d[5]
        f_rec = parse_fecha(d[7])
        f_ent = parse_fecha(d[8])
        ubic_rec = str(d[9] or "").strip()
        ubic_ent = str(d[11] or "").strip()
        tarifa = d[12]
        agencia = str(d[15] or "").strip()

        # Col A: Letra
        primer_char = (nombre or "").strip()[0].upper() if (nombre or "").strip() else ""
        letra_actual = primer_char if primer_char and "A" <= primer_char <= "Z" else ""
        if letra_actual and letra_actual != letra_anterior:
            letra_val = letra_actual
            letra_anterior = letra_actual
        else:
            letra_val = ""
            letra_actual = letra_actual  # keep tracking

        # Col C: Marca
        marca_val = marca_nombre

        # Col D: **
        col_d_val = calcular_columna_d(flag_col_a, plus_col_c, letra_marca)

        # Col G: **MAIN VIP**
        col_g_val, vip_tipo = calcular_main_vip(ubic_rec, agencia)

        # Escribir valores
        ws.cell(row=fila, column=1, value=letra_val)
        ws.cell(row=fila, column=2, value=nro_reserva)
        ws.cell(row=fila, column=3, value=marca_val)
        ws.cell(row=fila, column=4, value=col_d_val)
        ws.cell(row=fila, column=5, value=nombre)
        ws.cell(row=fila, column=6, value=clase)
        ws.cell(row=fila, column=7, value=col_g_val)
        ws.cell(row=fila, column=8, value=f_rec)
        ws.cell(row=fila, column=9, value=f_ent)
        ws.cell(row=fila, column=10, value=ubic_rec)
        ws.cell(row=fila, column=11, value=ubic_ent)
        ws.cell(row=fila, column=12, value=tarifa)
        ws.cell(row=fila, column=13, value=agencia)

        # ── Formato base ──
        for col in range(1, 14):
            c = ws.cell(row=fila, column=col)
            c.font = FONT_BOLD
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER
        ws.row_dimensions[fila].height = 18

        # Col A: Letra → rojo, fondo amarillo
        if letra_val:
            ws.cell(row=fila, column=1).font = Font(name="Calibri", bold=True, size=11, color=COLOR_ROJO)
            ws.cell(row=fila, column=1).fill = FILL_YELLOW

        # Col C: Marca → color según marca, negrita
        ws.cell(row=fila, column=3).font = Font(name="Calibri", bold=True, size=10, color=color_hex)

        # Col D: ** → rojo
        if col_d_val:
            ws.cell(row=fila, column=4).font = Font(name="Calibri", bold=True, size=10, color=COLOR_ROJO)

        # Col E: Nombre → color según marca
        ws.cell(row=fila, column=5).font = Font(name="Calibri", bold=True, size=10, color=color_hex)

        # Col G: MAIN VIP → fondo
        if vip_tipo == "vip":
            ws.cell(row=fila, column=7).fill = FILL_GREEN
        elif vip_tipo == "expedia":
            ws.cell(row=fila, column=7).fill = FILL_RED_LIGHT
        elif vip_tipo == "both":
            ws.cell(row=fila, column=7).fill = PatternFill("solid", fgColor="FFD966")

        # Borde superior en cambio de letra
        if letra_val:
            for col in range(1, 14):
                ws.cell(row=fila, column=col).border = Border(
                    left=Side(style="thin", color="AAAAAA"),
                    right=Side(style="thin", color="AAAAAA"),
                    top=Side(style="medium", color=COLOR_NEGRO),
                    bottom=Side(style="thin", color="AAAAAA"),
                )

        fila += 1

    # Congelar panel
    ws.freeze_panes = "A2"

    return fila - 2  # número de filas de datos


def process(file_in, file_out="GamePlan_Procesado.xlsx"):
    """Procesa el archivo Excel de manifiesto y genera el reporte."""
    wb_origen = load_workbook(file_in, data_only=True)
    ws_origen = wb_origen.active

    # Leer todas las filas
    datos_raw = []
    for row in ws_origen.iter_rows(min_row=2, values_only=True):
        if len(row) < 18:
            continue
        # row[0]=A, [1]=B, [2]=C, [3]=D, [4]=E, [5]=F, [6]=G, [7]=H, [8]=I,
        # [9]=J, [10]=K, [11]=L, [12]=M, [13]=N, [14]=O, [15]=P, [16]=Q, [17]=R
        ubic = str(row[9] or "").strip()
        if not ubic:
            continue
        marca_nombre, letra_marca, color_hex = determinar_marca(ubic)
        if not marca_nombre:
            continue

        # Guardar fila: 16 campos + marca derivada
        datos_raw.append({
            "ubic": ubic,
            "marca": marca_nombre,
            "letra_marca": letra_marca,
            "color": color_hex,
            "datos": list(row[:18]),  # hasta col R (índice 17)
        })

    if not datos_raw:
        print("No se encontraron datos válidos en el archivo.")
        return

    # Clasificar por marca
    marcas_map = {
        "ALAMO/": [],
        "ENTERPRISE/": [],
        "NATIONAL/": [],
    }
    colores_map = {
        "ALAMO/": COLOR_AZUL,
        "ENTERPRISE/": COLOR_NEGRO,
        "NATIONAL/": COLOR_VERDE,
    }
    letras_map = {
        "ALAMO/": "I",
        "ENTERPRISE/": "P",
        "NATIONAL/": "E",
    }

    for item in datos_raw:
        m = item["marca"]
        if m in marcas_map:
            # datos: [col_a(flag), col_b, col_c(plus), col_d, col_e, col_f, col_g,
            #         col_h, col_i, col_j, col_k, col_m, col_r(agencia),
            #         flag_col_a_orig, plus_col_c_orig]
            d = item["datos"]
            registro = [
                d[0],   # col A (flag ~/*)
                d[1],   # col B (N° reserva)
                d[2],   # col C (+)
                d[3],   # col D (nombre)
                d[3],   # col E = D (nombre)
                d[5],   # col F (clase)
                d[6],   # col G
                d[7],   # col H (fecha rec)
                d[8],   # col I (fecha ent)
                d[9],   # col J (ubic rec)
                d[10],  # col K (ubic ent)
                d[11],  # col L
                d[12],  # col M (tarifa)
                d[0],   # flag col A original
                d[2],   # plus col C original
                d[17],  # col R (agencia)
            ]
            marcas_map[m].append(registro)

    # Crear libro de salida
    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    total = 0
    for marca_nombre in ["ALAMO/", "ENTERPRISE/", "NATIONAL/"]:
        datos = marcas_map[marca_nombre]
        if not datos:
            continue
        nombre_hoja = marca_nombre.replace("/", "")
        ws = wb.create_sheet(title=nombre_hoja)
        n = escribir_hoja(ws, datos, marca_nombre,
                           letras_map[marca_nombre],
                           colores_map[marca_nombre])
        total += n
        print(f"  {nombre_hoja}: {n} registros")

    wb.save(file_out)
    print(f"\n✅ Reporte guardado: {file_out} ({total} registros totales)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 gameplan_reservas.py manifiesto.xlsx [salida.xlsx]")
        sys.exit(1)
    archivo_in = sys.argv[1]
    archivo_out = sys.argv[2] if len(sys.argv) > 2 else "GamePlan_Procesado.xlsx"
    process(archivo_in, archivo_out)
